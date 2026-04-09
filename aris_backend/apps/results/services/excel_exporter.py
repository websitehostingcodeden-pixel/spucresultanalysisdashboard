"""
STRICT EXCEL EXPORT ENGINE

Generates presentation-ready Excel files with exact formatting.

Sheets:
1. College Toppers
2. Science Toppers
3. Commerce Toppers
4. Section Performance
5. Subject Analysis
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

from apps.results.models import StudentResult, AnalyticsSnapshot


class ExcelExporter:
    """Generate Excel exports from analytics snapshots"""

    HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    SUBHEADER_FILL = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
    RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

    @staticmethod
    def _style_header(ws, row_num, columns):
        """Apply header styling to a row"""
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.border = ExcelExporter.BORDER
            cell.alignment = ExcelExporter.CENTER_ALIGN

    @staticmethod
    def _style_row(ws, row_num, columns_count, fill_color=None):
        """Apply row styling"""
        for col_num in range(1, columns_count + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = ExcelExporter.BORDER
            cell.alignment = ExcelExporter.LEFT_ALIGN

    @staticmethod
    def _auto_width(ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _get_analytics(upload_id):
        """Fetch analytics snapshot safely"""
        try:
            snapshot = AnalyticsSnapshot.objects.get(
                upload_log_id=upload_id,
                scope="UPLOAD"
            )
            if snapshot.is_expired():
                return None
            return snapshot.analytics_data
        except AnalyticsSnapshot.DoesNotExist:
            return None

    @staticmethod
    def _get_toppers(analytics_data, stream_filter=None):
        """Extract toppers from analytics"""
        if not analytics_data or "toppers" not in analytics_data:
            return []
        
        toppers = analytics_data["toppers"]
        if stream_filter:
            toppers = [t for t in toppers if t.get("stream") == stream_filter]
        
        # Sort by total marks descending, then by reg_no
        toppers = sorted(
            toppers,
            key=lambda x: (-x.get("grand_total", 0), x.get("reg_no", ""))
        )
        
        return toppers

    @staticmethod
    def _get_stream_performance(analytics_data):
        """Extract stream performance summary"""
        if not analytics_data or "stream_summary" not in analytics_data:
            return {}
        return analytics_data["stream_summary"]

    @staticmethod
    def _get_section_performance(analytics_data):
        """Extract section performance summary"""
        if not analytics_data or "section_summary" not in analytics_data:
            return {}
        return analytics_data["section_summary"]

    @staticmethod
    def _get_subject_analysis(analytics_data):
        """Extract subject analysis"""
        if not analytics_data or "subject_analysis" not in analytics_data:
            return {}
        return analytics_data["subject_analysis"]

    @classmethod
    def create_college_toppers_sheet(cls, wb, analytics_data):
        """Create sheet: College Toppers (Top 10 across all streams)"""
        ws = wb.create_sheet("College Toppers", 0)
        
        toppers = cls._get_toppers(analytics_data)[:10]
        
        headers = ["Rank", "Registration", "Stream", "Section", "Total Marks", "Percentage", "Grade"]
        cls._style_header(ws, 1, headers)
        
        for idx, topper in enumerate(toppers, 1):
            row = idx + 1
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = topper.get("reg_no", "")
            ws.cell(row=row, column=3).value = topper.get("stream", "")
            ws.cell(row=row, column=4).value = topper.get("section", "")
            ws.cell(row=row, column=5).value = topper.get("grand_total", 0)
            ws.cell(row=row, column=6).value = f"{topper.get('percentage', 0):.2f}%"
            ws.cell(row=row, column=7).value = topper.get("result_class", "")
            
            cls._style_row(ws, row, len(headers))
        
        # Alternate row colors
        for row in range(2, len(toppers) + 2):
            if row % 2 == 0:
                cls._style_row(ws, row, len(headers), fill_color="E8F0F8")
        
        cls._auto_width(ws)

    @classmethod
    def create_science_toppers_sheet(cls, wb, analytics_data):
        """Create sheet: Science Toppers (Top 10 SCIENCE stream)"""
        ws = wb.create_sheet("Science Toppers", 1)
        
        toppers = cls._get_toppers(analytics_data, stream_filter="SCIENCE")[:10]
        
        headers = ["Rank", "Registration", "Section", "Total Marks", "Percentage", "Grade"]
        cls._style_header(ws, 1, headers)
        
        for idx, topper in enumerate(toppers, 1):
            row = idx + 1
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = topper.get("reg_no", "")
            ws.cell(row=row, column=3).value = topper.get("section", "")
            ws.cell(row=row, column=4).value = topper.get("grand_total", 0)
            ws.cell(row=row, column=5).value = f"{topper.get('percentage', 0):.2f}%"
            ws.cell(row=row, column=6).value = topper.get("result_class", "")
            
            cls._style_row(ws, row, len(headers))
        
        for row in range(2, len(toppers) + 2):
            if row % 2 == 0:
                cls._style_row(ws, row, len(headers), fill_color="E8F0F8")
        
        cls._auto_width(ws)

    @classmethod
    def create_commerce_toppers_sheet(cls, wb, analytics_data):
        """Create sheet: Commerce Toppers (Top 10 COMMERCE stream)"""
        ws = wb.create_sheet("Commerce Toppers", 2)
        
        toppers = cls._get_toppers(analytics_data, stream_filter="COMMERCE")[:10]
        
        headers = ["Rank", "Registration", "Section", "Total Marks", "Percentage", "Grade"]
        cls._style_header(ws, 1, headers)
        
        for idx, topper in enumerate(toppers, 1):
            row = idx + 1
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = topper.get("reg_no", "")
            ws.cell(row=row, column=3).value = topper.get("section", "")
            ws.cell(row=row, column=4).value = topper.get("grand_total", 0)
            ws.cell(row=row, column=5).value = f"{topper.get('percentage', 0):.2f}%"
            ws.cell(row=row, column=6).value = topper.get("result_class", "")
            
            cls._style_row(ws, row, len(headers))
        
        for row in range(2, len(toppers) + 2):
            if row % 2 == 0:
                cls._style_row(ws, row, len(headers), fill_color="E8F0F8")
        
        cls._auto_width(ws)

    @classmethod
    def create_section_performance_sheet(cls, wb, analytics_data):
        """Create sheet: Section Performance"""
        ws = wb.create_sheet("Section Performance", 3)
        
        section_data = cls._get_section_performance(analytics_data)
        
        headers = ["Section", "Total Students", "Average Marks", "Distinction", "First Class", "Second Class", "Pass", "Fail"]
        cls._style_header(ws, 1, headers)
        
        row = 2
        for section, data in sorted(section_data.items()):
            ws.cell(row=row, column=1).value = section
            ws.cell(row=row, column=2).value = data.get("total_students", 0)
            ws.cell(row=row, column=3).value = f"{data.get('average_marks', 0):.2f}"
            ws.cell(row=row, column=4).value = data.get("grade_distribution", {}).get("DISTINCTION", 0)
            ws.cell(row=row, column=5).value = data.get("grade_distribution", {}).get("FIRST_CLASS", 0)
            ws.cell(row=row, column=6).value = data.get("grade_distribution", {}).get("SECOND_CLASS", 0)
            ws.cell(row=row, column=7).value = data.get("grade_distribution", {}).get("PASS", 0)
            ws.cell(row=row, column=8).value = data.get("grade_distribution", {}).get("FAIL", 0)
            
            cls._style_row(ws, row, len(headers))
            if row % 2 == 0:
                cls._style_row(ws, row, len(headers), fill_color="E8F0F8")
            
            row += 1
        
        cls._auto_width(ws)

    @classmethod
    def create_subject_analysis_sheet(cls, wb, analytics_data):
        """Create sheet: Subject Analysis"""
        ws = wb.create_sheet("Subject Analysis", 4)
        
        subject_data = cls._get_subject_analysis(analytics_data)
        
        headers = ["Subject", "Total Students", "Average Score", "Max Score", "Min Score", "Pass Rate"]
        cls._style_header(ws, 1, headers)
        
        row = 2
        for subject, data in sorted(subject_data.items()):
            ws.cell(row=row, column=1).value = subject
            ws.cell(row=row, column=2).value = data.get("total_students", 0)
            ws.cell(row=row, column=3).value = f"{data.get('average_score', 0):.2f}"
            ws.cell(row=row, column=4).value = data.get("max_score", 0)
            ws.cell(row=row, column=5).value = data.get("min_score", 0)
            ws.cell(row=row, column=6).value = f"{data.get('pass_rate', 0):.2f}%"
            
            cls._style_row(ws, row, len(headers))
            if row % 2 == 0:
                cls._style_row(ws, row, len(headers), fill_color="E8F0F8")
            
            row += 1
        
        cls._auto_width(ws)

    @classmethod
    def generate_export(cls, upload_id):
        """
        Generate complete Excel export with all 5 sheets
        
        Returns:
            BytesIO object with Excel file
        
        Raises:
            ValueError if analytics not found
        """
        # Fetch analytics snapshot
        analytics_data = cls._get_analytics(upload_id)
        if not analytics_data:
            raise ValueError(f"Analytics snapshot not found for upload {upload_id}")
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create all sheets
        cls.create_college_toppers_sheet(wb, analytics_data)
        cls.create_science_toppers_sheet(wb, analytics_data)
        cls.create_commerce_toppers_sheet(wb, analytics_data)
        cls.create_section_performance_sheet(wb, analytics_data)
        cls.create_subject_analysis_sheet(wb, analytics_data)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

    @classmethod
    def export_to_file(cls, upload_id, filename):
        """Export to file on disk"""
        output = cls.generate_export(upload_id)
        with open(filename, 'wb') as f:
            f.write(output.getvalue())
        return filename
